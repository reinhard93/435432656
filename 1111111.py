import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# 打开 Excel 文件
workbook_wallet = openpyxl.load_workbook('./wallet.xlsx')
# 获取默认的工作表
worksheet = workbook_wallet.active
# 获取第二行数据
row = 2  # 第二行的索引为2
Address = worksheet.cell(row=row, column=1).value
Seed_Phrase = worksheet.cell(row=row, column=2).value
PrivateKey = worksheet.cell(row=row, column=3).value
print(Address, Seed_Phrase, PrivateKey)
# 创建 ChromeOptions 对象
options = Options()
# 添加插件路径
plugin_path = './metamask.crx'
options.add_extension(plugin_path)
# 设置 Chrome WebDriver 的路径
chromedriver_path = '/path/to/chromedriver'
# 创建 Chrome WebDriver，并传递 options 参数
driver = webdriver.Chrome(options=options)
driver.set_window_size(1280, 900)
# 等待出现2个窗口，等待时间为10秒
WebDriverWait(driver, 100).until(EC.number_of_windows_to_be(2))
# 通过窗口句柄中，检测窗口1标题是否为metamask
metamask_handle = driver.window_handles[1]
# 切换metamask窗口
driver.switch_to.window(metamask_handle)
# 判断窗口url中是否含有home文本，等待时间为5秒
WebDriverWait(driver, 5).until(EC.url_contains('home'))
# 等待账户输入元素
el_1 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="onboarding__terms-checkbox"]'))
el_1.click()
el_2 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="app-content"]/div/div[2]/div/div/div/ul/li[3]/button'))
el_2.click()
el_3 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH,
                                                                         '//*[@id="app-content"]/div/div[2]/div/div/div/div/button[1]'))
el_3.click()
seed_phrase_words = Seed_Phrase.split()
# 定位助记词输入框并填入助记词
for index, word in enumerate(seed_phrase_words):
    input_xpath_1 = f'//*[@id="import-srp__srp-word-{index}"]'
    input_element = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_1))
    input_element.send_keys(word)
# 定位并点击指定的元素
el_3 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[4]/div/button'
button_element = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, el_3))
button_element.click()
# 输入密码导入
mm = 'aa123456'
input_xpath_2 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[1]/label/input'
el_4 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_2))
el_4.send_keys(mm)
# 确认密码
input_xpath_3 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[2]/label/input'
el_5 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_3))
el_5.send_keys(mm)
# 打钩
input_xpath_4 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/div[3]/label/input'
el_6 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_4))
el_6.click()
# 导入我的钱包
input_xpath_5 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/form/button'
el_7 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_5))
el_7.click()
input_xpath_6 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_8 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_6))
el_8.click()
input_xpath_7 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_9 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_7))
el_9.click()
input_xpath_8 = '//*[@id="app-content"]/div/div[2]/div/div/div/div[2]/button'
el_10 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_8))
el_10.click()
input_xpath_9 = '//*[@id="popover-content"]/div/div/section/div[2]/div/button/span'
el_9 = WebDriverWait(driver, timeout=100).until(lambda d: d.find_element(By.XPATH, input_xpath_99))
el_9.click()
time.sleep(10000)


# 读取 xlsx 文件
workbook_tw = openpyxl.load_workbook('./tw_accounts.xlsx')
# 获取默认的工作表
worksheet = workbook_wallet.active
# 获取第二行数据
row = 2  # 第二行的索引为2
username = worksheet.cell(row=row, column=1).value
password  = worksheet.cell(row=row, column=2).value
token  = worksheet.cell(row=row, column=3).value
print(username, password, token)